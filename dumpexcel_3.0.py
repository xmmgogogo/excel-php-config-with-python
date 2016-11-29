#-*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook
import time
import re
import json
import os
import sys
import paramiko
from ftplib import FTP

#更新svn
print "稍等>>>>>>>>>>>>>>> SVN更新中"
os.system("svn update D:\TD_Doc\TD2工作目录\配置文档")

# excel文件名
excelMaps = [
    'Task_VA', 'newbie_VA', 'TD_Achieves', 'lang_VA', 'HERO_VA', 'Home_VA', 'Item_VA', 'Monsters_VA',
    'SKILL_VA', 'TD_WareHouse', 'WAVES_VA', 'SUB_WAVES_VA', 'Task-test', 'TOWER_VA', 'tollgates_VA',
    'Tavern_VA', 'TD_Hire', 'Soldiers_VA', 'endlessWaves', 'Orders_VA', 'Reward_VA', 'Train_VA',
    'TD_SpiritInfo', 'Endless_Sub_Waves', 'barracks_VA', 'Activity_VA', 'wildMonster_VA', 'TD_explosion',
    'TD_Expedition', 'pop', 'PVP_VA', 'NPC_VA', 'Vip_VA', 'TD_Technology', 'TD_Competition', 'Dock_VA',
    'gameSetconfig', 'dropInfo_VA','Port_VA','Dragonester_VA','Totem_VA','HerrCountry_VA','Ornament_VA',
	'Union_VA'
]

# excelMaps = ['Task', 'newbie', 'TD_Achieves', 'lang', 'TD_HERO', 'TD_Home', 'TD_Item', 'TD_Monsters', 'TD_SKILL', 'TD_WareHouse', 'WAVES', 'SUB_WAVES', 'Task-test', 'TD_TOWER', 'tollgates', 'TD_Tavern', 'TD_Hire', 'TD_Soldiers', 'endlessWaves', 'TD_Orders', 'TD_Reward', 'TD_Train', 'TD_SpiritInfo','Endless_Sub_Waves','TD_barracks','TD_Activity','wildMonster','TD_explosion','TD_Expedition','pop', 'TD_PVP', 'TD_NPC', 'TD_Vip', 'TD_Technology', 'TD_Competition', 'TD_Dock'];

#过滤的sheet
fitter_keys = ['levelUp', '', 'game_init', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']
#传表的文件夹区分（通用配置和平台单独配置）
common_keys = ['worldBoss', 'gameSetconfig', 'competitionProgress', 'orders', 'unlock']
#特殊配置文件
special_keys = ['gameSetconfig']

def usage():
    print "1，命令行运行c:\dumpexcel_2.0.py"
    print "2，选择需要导出的Excel"
    print "3，选择是否上传到服务器"
    print "4，选择上传到哪一个服务器"


def main():
    #格式：0,1,2|1|dev_branch
    #默认从001里面获取?

    rName = '''
    #########################新版改动--->1,手动输入excell目录 2,指定上传环境#########################
    请【选择】文件序号(3,15,21)|上传版本(2.0, ArmorGames)|上传环境(branch, cehua, master)|是否上传(1)
    参考样式:4,7|1|2.0|branch(默认branch)     【TD1 - kot】
    参考样式:3,2|2|EN|release(默认release)    【TD2 - vikingage】

    ////////////////////////////////////////////////////////////////////////////////////////////////////////////
    / 0   = Task      | 1   = Newbie        | 2   = Achieves     | 3   = lang               | 4  = Hero        /
    / 5   = Home      | 6   = Item          | 7   = Monsters     | 8   = Skill              | 9  = WareHouse   /
    / 10  = Waves     | 11  = Sub_waves     | 12  = Task-test    | 13  = Tower              | 14 = Tollgates   /
    / 15  = Tavern    | 16  = Hire          | 17  = Soldiers     | 18  = EndlessWaves       | 19 = Orders      /
    / 20  = Reward    | 21  = Train         | 22  = SpiritInfo   | 23  = Endless_Sub_Waves  | 24 = Barracks    /
    / 25  = Activity  | 26  = WildMonster   | 27  = Explosion    | 28  = Expedition         | 29 = Pop         /
    / 30  = PVP       | 31  = NPC           | 32  = Vip          | 33  = Technology         | 34 = Competition /
    / 35  = Dock      | 36  = gameSetconfig | 37  = dropInfos    | 38  = Port               | 39 = Dragonester /
    / 40  = Totem     | 41  = HerrCountry   | 42  = Ornament     | 43  = Union_VA           | 999 = (全部上传) /
    ////////////////////////////////////////////////////////////////////////////////////////////////////////////
    '''
    rawSomething = raw_input(rName)
    rawSomething = rawSomething.split('|')
    if len(rawSomething) < 1:
        print "传入序号不正确！重新传入！"
        usage()
        sys.exit()

    #start
    startTime = time.time()

    #['0,1,2', '2.0', 'dev_branch', 1]
    #导出name
    excelNames = rawSomething[0].split(',')
    if '999' in excelNames:
        excelNames = range(0, 36)
        del excelNames[12]

    # 默认上传
    isUpload = 1
    #选择版本目录
    workSpace = 'EN'
    # 默认上传环境
    baseUploadDev = 'branch'
    # 上传版本
    cur_version = 2

    if len(rawSomething) >= 2:
        if rawSomething[1]:
            cur_version = int(rawSomething[1])
    if len(rawSomething) >= 3:
        if rawSomething[2]:
            workSpace = rawSomething[2]
    if len(rawSomething) >= 4:
        if rawSomething[3]:
            baseUploadDev = rawSomething[3]

    if cur_version == 1:
        baseUploadDev = 'dev_' + baseUploadDev
        fileConfig = json.load((file('c:\excelConfig.json')))
    else:
        if baseUploadDev == 'branch':#默认
            baseUploadDev = 'release'
        fileConfig = json.load((file('c:\excelConfig-vikingage.json')))

    filePath = fileConfig['filePath'] % workSpace
    # print type(cur_version)
    # print filePath

    #遍历全部导出excel
    for getExcelNameId in excelNames:
        getExcelNameId = int(getExcelNameId)

        #加入判断条件
        if excelMaps[getExcelNameId] not in special_keys:
            isUploadFlag = 0

            #读取excel2007文件
            filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
            if os.path.isfile(filePathName) == False:
                filePathName = filePath + 'TD_' + excelMaps[getExcelNameId] + '.xlsx'
                if os.path.isfile(filePathName) == False:
                    continue

            readExcel = load_workbook(filePathName)

            #取第一张表
            sheetNames = readExcel.get_sheet_names()
            savePath = fileConfig['savePath']
            for sheetName in sheetNames:
                # 过滤不需要生成配置文件的excel名
                if re.match('#', sheetName) or sheetName in fitter_keys:
                    continue

                ws = readExcel.get_sheet_by_name(sheetName)
                columnKey = {}
                allData = {}
                keyData = []

                for _row in range(ws.get_highest_row()):
                    # 将每行第一列作为key
                    firstCellVal = ws.cell(row=_row, column=0).value
                    # 申明新的字典，保存数据
                    if _row > 2:
                        allData[firstCellVal] = {}

                    for _column in range(ws.get_highest_column()):
                        # 获取数据集固定的字段名
                        firstCellKey = ws.cell(row=2, column=_column).value

                        if not firstCellKey:
                            continue
                        if type(firstCellKey) != unicode:
                            firstCellKey = str(firstCellKey)
                            # 获取每行每列的字段数值
                        cellVal = ws.cell(row=_row, column=_column).value

                        # print cellVal, _row, _column #检测报错行列

                        # 第二行存储字段名
                        if _row == 2:
                            if not re.search('\*', firstCellKey):
                                if firstCellKey not in keyData:
                                    keyData.append(firstCellKey)
                                columnKey[_column] = cellVal

                        # 从第二行开始
                        if _row > 2:
                            if re.search('\*', firstCellKey):
                                continue

                            try:
                                cellKey = columnKey[_column]
                                if cellVal == None:
                                    cellVal = ''
                                allData[firstCellVal][cellKey] = cellVal
                            except:
                                print '=====error======:', _column
                                raise

                f = open(savePath + sheetName + '.php', 'w+')
                f.write(json.dumps(allData))
                f = open(savePath + sheetName + '.txt', 'w+')
                f.write(json.dumps(keyData))
                f.close()

                os.system('php c:\\phpDecode.php ' + sheetName + '.php ' + str(cur_version))

                #这里多一个操作，若属于特殊分表文件，进行分表
                if getExcelNameId in [6, 11, 10, 14]:
                    # print getExcelNameId , "==================================" , isUploadFlag
                    os.system('php c:\\cuttingConfig.php ' + sheetName + ' ' + str(cur_version))
                    if getExcelNameId in [11, 10, 14] and isUploadFlag == 0:
                        # print baseUploadDev
                        ftpClass('up', 'wavesInfo', 'wavesInfo', baseUploadDev, fileConfig, cur_version)
                        isUploadFlag = 1

                #ftp
                if isUpload:
                    ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig, cur_version)

        else:
            filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
            # print filePathName
            readExcel = load_workbook(filePathName)

            sheetNames = readExcel.get_sheet_names()
            savePath = fileConfig['savePath']
            for sheetName in sheetNames:
                if re.match('#', sheetName) or sheetName in fitter_keys:
                    continue

                ws = readExcel.get_sheet_by_name(sheetName)
                allData = {}
                keyData = []

                for _row in range(ws.get_highest_row()):
                    if _row > 0:
                        firstCellVal = ws.cell(row=_row, column=0).value
                        allData[firstCellVal] = {}

                        allData[firstCellVal] = [ws.cell(row=_row, column=1).value, ws.cell(row=_row, column=2).value]
                        keyData.append(firstCellVal)

                f = open(savePath + sheetName + '.php', 'w+')
                f.write(json.dumps(allData))
                f = open(savePath + sheetName + '.txt', 'w+')
                f.write(json.dumps(keyData))
                f.close()

                os.system('php c:\\phpDecode-gameSet.php ' + sheetName + '.php ' + str(cur_version))

                #ftp
                if isUpload:
                    ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig, cur_version)

                    # if isUpload == 0:
                    #     os.system('c:\\copyGit_Svn.py ')

    print "\n耗时:", str("%.2f" % (time.time() - startTime)) + "s."

# FTP文件操作
def ftpClass(type, baseFile, serverFile, devName, fileConfig, cur_version):
    host = fileConfig['server']['host']
    user = fileConfig['server']['user']
    pwd = fileConfig['server']['pwd']

    # config name ----> see in xml
    #这里比较坑爹，Unicode和string 混一起必然报错.一定先将Unicode 转换为string
    configPath = fileConfig['configName'].encode("utf-8")

    if cur_version == 1:
        ftp = FTP()
        ftp.connect(host)
        ftp.login(user, pwd)
    else:
        # ssh_connect(host, 22, user, pwd)
        # client = paramiko.Transport((host, 22))
        # client.connect(username=user, password=pwd)
        # ftp = paramiko.SFTPClient.from_transport(client)
        ftp = FTP()
        ftp.connect(host)
        ftp.login(user, pwd)

        # 特殊处理viking-age
        if baseFile not in common_keys:
            configPath = 'commonConfig'
            # print configPath, baseFile

    # 这里写死为1，因为强制让运维修改为ftp而非使用sftp (去掉注释可兼容sftp)
    cur_version = 1

    basePath = fileConfig['basePath']
    serverPath = fileConfig['serverPath']
    serverPath = serverPath % (devName, configPath)

    basePath = basePath % configPath
    baseFile = basePath + '/' + baseFile
    serverFile = serverPath + '/' + serverFile
    # print baseFile, os.path.isdir(baseFile)

    if type == 'up':
        if os.path.isfile(baseFile):
            ftpUp(ftp, baseFile, serverFile, configPath, cur_version)
        if os.path.isdir(baseFile):
            fNames = os.listdir(baseFile)

            for fName in fNames:
                if fName != '.svn':
                    ftpUp(ftp, baseFile + '/' + fName, serverFile + '/' + fName, devName, configPath, cur_version)

    if type == 'down':
        ftpDown(ftp, baseFile, serverFile, cur_version)

    if cur_version == 1:
        ftp.close()
    else:
        client.close()

# FTP文件上传
def ftpUp(ftp, baseFile, serverFile, devName, configPath, cur_version):
    # print baseFile, serverFile
    if cur_version == 1:
        f = open(baseFile, 'r')
        result = ftp.storlines('STOR %s' % ('/' + serverFile), f)
        f.close()
    else:
        result = ftp.put(baseFile, serverFile)

    # print type(result)
    print "成功上传至-[" + baseFile.encode("utf-8")[-9:] + "]->" + devName + "(" + configPath + ")"

# FTP文件下载
def ftpDown(ftp, baseFile, serverFile, cur_version):
    if cur_version == 1:
        f = open(baseFile, 'wb')
        result = ftp.retrbinary('RETR %s' % ('/' + serverFile), f.write)
        f.close()
    else:
        result = ftp.get(baseFile, serverFile)

    print "成功下载-[" + baseFile.encode("utf-8")[-9:] + "]->" + devName + "(" + configPath + ")"


def ssh_connect(server_ip, server_port, server_user, server_passwd):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(server_ip, server_port, server_user, server_passwd)
    return ssh


def ssh_disconnect(client):
    client.close()


def win_to_linux(sftp, localpath, remotepath):
    '''
    windows向linux服务器上传文件.
    localpath  为本地文件的绝对路径。如：D:\test.py
    remotepath 为服务器端存放上传文件的绝对路径,而不是一个目录。如：/tmp/my_file.txt
    '''
    sftp.put(localpath, remotepath)


def linux_to_win(sftp, localpath, remotepath):
    '''
    从linux服务器下载文件到本地
    localpath  为本地文件的绝对路径。如：D:\test.py
    remotepath 为服务器端存放上传文件的绝对路径,而不是一个目录。如：/tmp/my_file.txt
    '''
    sftp.get(remotepath, localpath)


if __name__ == '__main__':
    opts = sys.argv[1:]
    for op in opts:
        print op
        if op:
            usage()
            sys.exit()

    main()