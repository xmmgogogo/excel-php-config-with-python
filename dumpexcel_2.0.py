#-*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook
import time
import re
import json
import os
import sys
from ftplib import FTP

#更新svn
print "稍等>>>>>>>>>>>>>>> SVN更新中"
# os.system("svn update D:\TD_Doc\数值")

FILE_CONFIG = 'excelConfig'#后2个没用了
# FILE_CONFIG = ['', 'excelConfig001', 'excelConfig002', 'excelConfig001', 'excelConfig001', 'excelConfig001']#后2个没用了
# excel文件名
excelMaps = ['Task', 'newbie', 'TD_Achieves', 'lang', 'TD_HERO', 'TD_Home', 'TD_Item', 'TD_Monsters', 'TD_SKILL', 'TD_WareHouse', 'WAVES', 'SUB_WAVES', 'Task-test', 'TD_TOWER', 'tollgates', 'TD_Tavern', 'TD_Hire', 'TD_Soldiers', 'endlessWaves', 'TD_Orders', 'TD_Reward', 'TD_Train', 'TD_SpiritInfo','Endless_Sub_Waves','TD_barracks','TD_Activity','wildMonster','TD_explosion','TD_Expedition','pop', 'TD_PVP', 'TD_NPC', 'TD_Vip', 'TD_Technology', 'TD_Competition', 'TD_Dock'];
#过滤的sheet
fitter_keys = ['levelUp', '', 'game_init', 'Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']#heroStrengthenOdd

def usage():
    print "1，命令行运行c:\dumpexcel_2.0.py"
    print "2，选择需要导出的Excel"
    print "3，选择是否上传到服务器"
    print "4，选择上传到哪一个服务器"

def main():
    #格式：0,1,2|1|dev_branch
    #默认从001里面获取

    rName = '''
    #########################新版改动--->1,手动输入excell目录 2,指定上传环境#########################
    请【选择】文件序号(3,15,21)|配置目录(2.0, ArmorGames)|上传环境(branch, cehua, master)|是否上传(1)
    参考样式：4,7,15|ArmorGames|branch|1。默认上传branch，只需要填写序号即可！

    //////////////////////////////////////////////////////////////////////////////////////////////////////////
    / 0   = Task      | 1   = Newbie      | 2   = Achieves     | 3   = lang               | 4  = Hero        /
    / 5   = Home      | 6   = Item        | 7   = Monsters     | 8   = Skill              | 9  = WareHouse   /
    / 10  = Waves     | 11  = Sub_waves   | 12  = Task-test    | 13  = Tower              | 14 = Tollgates   /
    / 15  = Tavern    | 16  = Hire        | 17  = Soldiers     | 18  = EndlessWaves       | 19 = Orders      /
    / 20  = Reward    | 21  = Train       | 22  = SpiritInfo   | 23  = Endless_Sub_Waves  | 24 = Barracks    /
    / 25  = Activity  | 26  = WildMonster | 27  = Explosion    | 28  = Expedition         | 29 = Pop         /
    / 30  = PVP       | 31  = NPC         | 32  = Vip          | 33  = Technology         | 34 = Competition / 
    / 35  = Dock      | 999 = (全部上传)  |                                                                  /
    //////////////////////////////////////////////////////////////////////////////////////////////////////////
    '''
    rawSomething = raw_input(rName)
    rawSomething = rawSomething.split('|')
    if len(rawSomething) < 1:
        print "传入序号不正确！重新传入！"
        usage()
        sys.exit()

    #['0,1,2', '2.0', 'dev_branch', 1]
    #导出name
    excelNames = rawSomething[0].split(',')
    if '999' in excelNames:
        excelNames = range(0, 35)
        del excelNames[12]

    #选择版本目录
    workSpace = '2.0'
    # 默认上传
    isUpload = 1
    # 默认上传环境
    baseUploadDev = 'branch'

    if len(rawSomething) >= 2:
        if rawSomething[1]:
            workSpace = rawSomething[1]
    if len(rawSomething) >= 3:
        if rawSomething[2]:
            baseUploadDev = rawSomething[2]
    if len(rawSomething) >= 4:
        if rawSomething[3]:
            isUpload = rawSomething[3]

    #add dev
    baseUploadDev = 'dev_' + baseUploadDev
    # print workSpace, baseUploadDev, isUpload

    #读取配置文件
    isUpload = int(isUpload)
    fileConfig = 'c:\\' + FILE_CONFIG + '.json'
    fileConfig = json.load((file(fileConfig)))

    filePath = fileConfig['filePath'] % workSpace
    # print filePath,excelNames;exit();
    
    isUploadFlag = 0

    for getExcelNameId in excelNames:
        getExcelNameId = int(getExcelNameId)
        # print eName
        #读取excel2007文件            
        filePathName = filePath + excelMaps[getExcelNameId] + '.xlsx'
        if os.path.isfile(filePathName) == False:
            filePathName = filePath + 'TD_' + excelMaps[getExcelNameId] + '.xlsx'
            if os.path.isfile(filePathName) == False:
                continue

        readExcel = load_workbook(filePathName)

        #取第一张表
        sheetNames = readExcel.get_sheet_names()
        savePath = fileConfig['savePath']
        for sheetName in sheetNames:
            # 过滤不需要生成配置文件的excel名
            if re.match('#', sheetName) or sheetName in fitter_keys:
                continue

            ws = readExcel.get_sheet_by_name(sheetName)
            columnKey = {}
            allData = {}
            keyData = []

            for _row in range(ws.get_highest_row()):
                # 将每行第一列作为key
                firstCellVal = ws.cell(row=_row, column=0).value
                # 申明新的字典，保存数据
                if _row > 2:
                    allData[firstCellVal] = {}

                for _column in range(ws.get_highest_column()):
                    # 获取数据集固定的字段名
                    firstCellKey = ws.cell(row=2, column=_column).value

                    if not firstCellKey:
                        continue
                    if type(firstCellKey) != unicode:
                        firstCellKey = str(firstCellKey)
                    # 获取每行每列的字段数值
                    cellVal = ws.cell(row=_row, column=_column).value

                    # print cellVal, _row, _column #检测报错行列

                    # 第二行存储字段名
                    if _row == 2:
                        if not re.search('\*', firstCellKey):
                            if firstCellKey not in keyData:
                                keyData.append(firstCellKey)
                            columnKey[_column] = cellVal

                    # 从第二行开始
                    if _row > 2:
                        if re.search('\*', firstCellKey):
                            continue

                        try:
                            cellKey = columnKey[_column]
                            if cellVal == None:
                                cellVal = ''
                            allData[firstCellVal][cellKey] = cellVal
                        except:
                            print '=====error======:',_column
                            raise

            f = open(savePath + sheetName + '.php', 'w+')
            f.write(json.dumps(allData))
            f = open(savePath + sheetName + '.txt', 'w+')
            f.write(json.dumps(keyData))
            f.close()

            os.system('php c:\\phpDecode.php ' + sheetName + '.php')

            #这里多一个操作，若属于特殊分表文件，进行分表
            if getExcelNameId in [6, 11, 10, 14]:
                # print getExcelNameId , "==================================" , isUploadFlag
                os.system('php c:\\cuttingConfig.php ' + sheetName)
                if getExcelNameId in [11, 10, 14] and isUploadFlag == 0:
                    # print baseUploadDev
                    ftpClass('up', 'wavesInfo', 'wavesInfo', baseUploadDev, fileConfig)
                    isUploadFlag = 1

            if isUpload:
                ftpClass('up', sheetName, sheetName, baseUploadDev, fileConfig)

    if isUpload == 0:
        os.system('c:\\copyGit_Svn.py ')

# FTP文件操作
def ftpClass(type, baseFile, serverFile, devName, fileConfig):
    host = fileConfig['server']['host']
    user = fileConfig['server']['user']
    pwd = fileConfig['server']['pwd']

    ftp = FTP()
    ftp.connect(host)
    ftp.login(user, pwd)

    # config name ----> see in xml
    #这里比较坑爹，Unicode和string 混一起必然报错.一定先将Unicode 转换为string
    configPath = fileConfig['configName'].encode("utf-8")
    # print configPath, devName

    basePath = fileConfig['basePath']
    serverPath = fileConfig['serverPath']
    serverPath = serverPath % (devName, configPath)

    basePath = basePath % configPath
    baseFile = basePath + '/' + baseFile
    serverFile = serverPath + '/' + serverFile
    # print baseFile, "||" , serverFile

    if type == 'up':
        if os.path.isfile(baseFile):
            ftpUp(ftp, baseFile, serverFile)
        if os.path.isdir(baseFile):
            fNames = os.listdir(baseFile)

            for fName in fNames:
                if fName != '.svn':
                    ftpUp(ftp, baseFile + '/' + fName, serverFile + '/' + fName, devName, configPath)

    if type == 'down':
        ftpDown(ftp, baseFile, serverFile)

    ftp.close()


# FTP文件上传
def ftpUp(ftp, baseFile, serverFile, devName, configPath):
    f = open(baseFile, 'r')
    # print baseFile, serverFile
    print "成功上传至-->" +devName +"("+configPath+ ")："+ ftp.storlines('STOR %s'% ('/' + serverFile), f)
    f.close()


# FTP文件下载
def ftpDown(ftp, baseFile, serverFile, devName):
    f = open(baseFile, 'wb')
    print "成功上传["+baseFile+"]至：" + ftp.retrbinary('RETR %s'% ('/' + serverFile), f.write)
    f.close()

if __name__ == '__main__':
    #开始时间
    startTime = time.time()

    opts = sys.argv[1:]
    for op in opts:
        print op
        if op:
            usage()
            sys.exit()

    main()
    print "\n耗时:",str("%.2f" % (time.time() - startTime)) + "s."