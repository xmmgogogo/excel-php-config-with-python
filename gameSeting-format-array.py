#-*- coding:utf-8 -*-
from openpyxl.reader.excel import load_workbook
import time
import re
import json
import os
import sys
from ftplib import FTP

fitter_keys = ['levelUp', 'game_init']
# rawSomething = raw_input('input name::')
# rawSomething = rawSomething.split('|')
# if len(rawSomething) < 1:
#     print "exit"
#     sys.exit()

workSpace = '2.0'
excelNames = ['TD_GameSetting']
isUpload = 1
baseUploadDev = 1
fileConfig = 'c:\\' + 'excelConfig' + '.json'
fileConfig = json.load((file(fileConfig)))
# print fileConfig

isUploadFlag = 0
for getExcelName in excelNames:
    filePath = fileConfig['filePath'] % workSpace
    filePathName = filePath + getExcelName + '.xlsx'
    # print filePathName
    readExcel = load_workbook(filePathName)

    sheetNames = readExcel.get_sheet_names()
    savePath = fileConfig['savePath']
    for sheetName in sheetNames:
        if re.match('#', sheetName) or sheetName in fitter_keys:
            continue

        ws = readExcel.get_sheet_by_name(sheetName)
        columnKey = {}
        allData = {}
        keyData = []

        for _row in range(ws.get_highest_row()):
            if _row > 0:
                firstCellVal = ws.cell(row=_row, column=0).value
                allData[firstCellVal] = {}

                # for _column in range(ws.get_highest_column()):
                #     if _column > 1:
                #         columnKey = ws.cell(row=_row, column=_column).value
                #         allData[firstCellVal] = columnKey

                allData[firstCellVal] = [ws.cell(row=_row, column=1).value, ws.cell(row=_row, column=2).value]
                keyData.append(firstCellVal)

        f = open(savePath + sheetName + '.php', 'w+')
        f.write(json.dumps(allData))
        f = open(savePath + sheetName + '.txt', 'w+')
        f.write(json.dumps(keyData))
        f.close()

        os.system('php c:\\phpDecode-gameSet.php ' + sheetName + '.php ')